import io
import logging
from datetime import timedelta

import openpyxl
import requests
import requests_cache
from django.core.management.base import BaseCommand
from django.db import connections, models, router, transaction
from django.utils.text import slugify
from openpyxl import load_workbook
from tqdm import tqdm

from charity_django.crie.models import (
    COMPANY_FORM_LOOKUP,
    Charity,
    CharityClassificationCategory,
    CharityFinancialYear,
    ClassificationTypes,
)
from charity_django.crie.utils import parse_classification, parse_classification_simple

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class Command(BaseCommand):
    help = "Import Charity Regulator Ireland data from an Excel file"
    page_size = 100_000
    base_url = "https://www.charitiesregulator.ie/media/1rviel2n/20251019-public-register-of-charities.xlsx"

    def _get_db(self):
        return router.db_for_write(Charity)

    def _get_connection(self):
        return connections[self._get_db()]

    def logger(self, message, error=False):
        if error:
            logger.error(message)
            return
        logger.info(message)

    def add_arguments(self, parser):
        parser.add_argument("--sample", type=int, default=0)
        parser.add_argument("--file", type=str, default=None)

    def handle(self, *args, **options):
        self.session = requests_cache.CachedSession(
            "demo_cache.sqlite",
            expire_after=timedelta(days=1),
            headers={"User-Agent": "charity-django-importer/1.0"},
        )
        self.sample = options.get("sample")
        self.file = options.get("file")

        self.classification_cache = {}
        self.charities = {}
        self.charity_fields = [
            f.get_attname_column()[1]
            for f in Charity._meta.fields
            if f.get_attname_column()[1] != "registered_charity_number"
        ]
        self.charity_account_fields = [
            f.get_attname_column()[1] for f in CharityFinancialYear._meta.fields
        ]

        self.names = set()

        self.fetch_file()

    def fetch_file(self) -> None:
        self.files = {}
        if self.file:
            self.logger(f"Using local file: {self.file}")
            with open(self.file, "rb") as f:
                file = io.BytesIO(f.read())
        else:
            self.logger(f"Downloading file from: {self.base_url}")
            try:
                r = self.session.get(self.base_url, verify=True)
            except requests.exceptions.SSLError:
                r = self.session.get(self.base_url, verify=False)
            r.raise_for_status()

            file = io.BytesIO(r.content)

        self.logger("Loading workbook...")
        wb = load_workbook(filename=file, read_only=True)

        # check that that sheets we're expecting are present
        PUBLIC_REGISTER_SHEET = "Public Register"
        ANNUAL_REPORTS_SHEET = "Annual Reports"
        expected_sheets = [PUBLIC_REGISTER_SHEET, ANNUAL_REPORTS_SHEET]
        for sheet in expected_sheets:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Expected sheet '{sheet}' not found in workbook")

        db = self._get_db()
        connection = self._get_connection()
        with connection.cursor(), transaction.atomic(using=db):
            classifications = set()
            classifications.update(
                self.get_charities_classifications(wb[PUBLIC_REGISTER_SHEET])
            )
            classifications.update(
                self.get_annual_reports_classifications(wb[ANNUAL_REPORTS_SHEET])
            )
            self.add_classification_categories(classifications)

            self.add_charities(wb[PUBLIC_REGISTER_SHEET])
            self.add_annual_reports(wb[ANNUAL_REPORTS_SHEET])
            self.update_latest_financial_years()

    def get_charities_classifications(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> set[tuple[str, ClassificationTypes]]:
        self.logger("Finding charity classification categories...")
        # import Public Register sheet
        headers = [cell.value.strip() if cell.value else None for cell in ws[2]]
        charity_classifications = set()
        for i, row in enumerate(
            tqdm(
                ws.iter_rows(min_row=3, values_only=True),
                desc="Processing charities",
                total=ws.max_row - 2,
            )
        ):
            if not any(row):
                break
            _, _, classifications = self.get_charity(dict(zip(headers, row)))
            charity_classifications.update(classifications)

        self.logger("Finished finding charity classification categories.")
        return charity_classifications

    def add_charities(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        self.logger("Importing charities...")
        # import Public Register sheet
        headers = [cell.value.strip() if cell.value else None for cell in ws[2]]
        for i, row in enumerate(
            tqdm(
                ws.iter_rows(min_row=3, values_only=True),
                desc="Processing charities",
                total=ws.max_row - 2,
            )
        ):
            if not any(row):
                break
            self.add_charity(dict(zip(headers, row)))
        self.logger("Finished importing charities.")

    def get_charity(
        self, record: dict
    ) -> tuple[dict, set[str], set[tuple[str, ClassificationTypes]]]:
        record = {
            slugify(k).replace("-", "_").replace("__", "_").replace("__", "_"): (
                v if v != "" else None
            )
            for k, v in record.items()
            if k and k != "Trustees (Start Date)"
        }

        # primary address field (replace multiple , , with single ,)
        if record.get("primary_address"):
            record["primary_address"] = ", ".join(
                [
                    line.strip()
                    for line in record["primary_address"].split(",")
                    if line.strip()
                ]
            )

        # governing form field
        if record.get("governing_form"):
            if record["governing_form"] not in COMPANY_FORM_LOOKUP:
                self.logger(
                    f"Unknown governing form: {record['governing_form']}", error=True
                )
            record["governing_form"] = COMPANY_FORM_LOOKUP.get(
                record["governing_form"], None
            )

        # name fields
        names = set([record["registered_charity_name"].strip()])
        if record.get("also_known_as"):
            for name in parse_classification_simple(record.pop("also_known_as")):
                names.add(name)

        classifications = set()

        #  classification primary/secondary sub field
        if record.get("charity_classification_primary_secondary_sub"):
            classification = record.pop("charity_classification_primary_secondary_sub")
            for c in parse_classification(classification):
                classifications.add(
                    (
                        c.strip("\"', "),
                        ClassificationTypes.CHARITY_CLASSIFICATION,
                    )
                )

        # also operates in field and charitable_purpose field
        fields = [
            ("also_operates_in", ClassificationTypes.OPERATES_IN),
            ("charitable_purpose", ClassificationTypes.CHARITABLE_PURPOSE),
        ]
        for field, classification_type in fields:
            if record.get(field):
                for c in parse_classification_simple(record.pop(field)):
                    classifications.add((c, classification_type))

        return record, names, classifications

    def add_charity(self, record: dict) -> None:
        record, names, classifications = self.get_charity(record)

        charity = Charity.objects.filter(
            registered_charity_number=record["registered_charity_number"]
        ).first()
        if not charity:
            charity = Charity(
                registered_charity_number=record["registered_charity_number"]
            )
        for key, value in record.items():
            if key in self.charity_fields:
                setattr(charity, key, value)
        charity.save()

        for name in names:
            charity.names.get_or_create(
                name=name,
                defaults={"language": "en"},
            )

        charity.classifications.set(
            [
                self._get_classification_category(c, classification_type)
                for c, classification_type in classifications
            ]
        )

    def get_annual_reports_classifications(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> set[tuple[str, ClassificationTypes]]:
        self.logger("Finding annual report classification categories...")
        # import Annual Reports sheet
        headers = [cell.value.strip() if cell.value else None for cell in ws[2]]
        charity_classifications = set()
        for i, row in enumerate(
            tqdm(
                ws.iter_rows(min_row=3, values_only=True),
                desc="Processing annual reports",
                total=ws.max_row - 2,
            )
        ):
            if not any(row):
                break
            _, classifications = self.get_annual_report(dict(zip(headers, row)))
            charity_classifications.update(classifications)
        self.logger("Finished finding annual report classification categories.")
        return charity_classifications

    def get_annual_report(
        self, record: dict
    ) -> tuple[dict, set[tuple[str, ClassificationTypes]]]:
        record = {
            slugify(k).replace("-", "_").replace("__", "_").replace("__", "_"): (
                v if v != "" else None
            )
            for k, v in record.items()
            if k and k != "Registered Charity Name"
        }
        record["charity_id"] = int(record.pop("registered_charity_number_rcn"))

        classifications = set()

        fields = [
            ("report_activity", ClassificationTypes.REPORT_ACTIVITY),
            ("beneficiaries", ClassificationTypes.BENEFICIARIES),
        ]
        for field, classification_type in fields:
            if record.get(field):
                for c in parse_classification_simple(record.pop(field)):
                    classifications.add((c, classification_type))
        return record, classifications

    def add_annual_report(self, record: dict) -> None:
        record, classifications = self.get_annual_report(record)
        charity_fy, created = CharityFinancialYear.objects.update_or_create(
            charity_id=record["charity_id"],
            period_end_date=record["period_end_date"].date(),
            defaults={
                key: value
                for key, value in record.items()
                if key in self.charity_account_fields
            },
        )
        charity_fy.classifications.set(
            [
                self._get_classification_category(c, classification_type)
                for c, classification_type in classifications
            ]
        )

    def add_classification_categories(
        self, classifications: set[tuple[str, ClassificationTypes]]
    ) -> None:
        self.logger("Adding classification categories...")
        for classification, classification_type in tqdm(
            classifications, desc="Processing classification categories"
        ):
            self._get_classification_category(classification, classification_type)
        self.logger("Finished adding classification categories.")

    def _get_classification_category(
        self, classification: str, classification_type: ClassificationTypes
    ) -> CharityClassificationCategory:
        key = (classification, classification_type)
        if key in self.classification_cache:
            return self.classification_cache[key]
        category = CharityClassificationCategory.objects.filter(
            classification_type=classification_type,
            classification_en=classification,
        ).first()
        if not category:
            category = CharityClassificationCategory.objects.filter(
                classification_type=classification_type,
                classification_ga=classification,
            ).first()
        if not category:
            category = CharityClassificationCategory.objects.create(
                classification_type=classification_type,
                classification_en=classification,
            )
        self.classification_cache[key] = category
        return category

    def add_annual_reports(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        self.logger("Importing annual reports...")
        # import Annual Reports sheet
        self.classifications: set[tuple[str, ClassificationTypes]] = set()

        def generate_annual_reports():
            headers = [cell.value.strip() if cell.value else None for cell in ws[2]]
            for i, row in enumerate(
                tqdm(
                    ws.iter_rows(min_row=3, values_only=True),
                    desc="Processing annual reports",
                    total=ws.max_row - 2,
                )
            ):
                if not any(row):
                    break
                record, classifications = self.get_annual_report(
                    dict(zip(headers, row))
                )
                for c in classifications:
                    self.classifications.add(
                        (
                            record["charity_id"],
                            record["period_end_date"].date(),
                            self._get_classification_category(c[0], c[1]),
                        )
                    )
                yield CharityFinancialYear(
                    **{
                        key: value
                        for key, value in record.items()
                        if key in self.charity_account_fields
                    }
                )

        CharityFinancialYear.objects.bulk_create(
            generate_annual_reports(),
            batch_size=self.page_size,
            update_conflicts=True,
            unique_fields=["charity_id", "period_end_date"],
            update_fields=[
                "period_start_date",
                "activity_description",
                "income_government_or_local_authorities",
                "income_other_public_bodies",
                "income_philantrophic_organisations",
                "income_donations",
                "income_trading_and_commercial_activities",
                "income_other_sources",
                "income_bequests",
                "gross_income",
                "gross_expenditure",
                "surplus_deficit_for_the_period",
                "cash_at_hand_and_in_bank",
                "other_assets",
                "total_assets",
                "total_liabilities",
                "net_assets_liabilities",
                "gross_income_schools",
                "gross_expenditure_schools",
                "number_of_employees",
                "number_of_full_time_employees",
                "number_of_part_time_employees",
                "number_of_volunteers",
            ],
        )

        self.logger("Finished importing annual reports.")

        self.logger("Importing annual report classifications...")
        self.logger("Fetching existing financial years...")
        lookup = {
            (cfy.charity_id, cfy.period_end_date): cfy.id
            for cfy in CharityFinancialYear.objects.all()
        }
        self.logger("Finished fetching existing financial years.")

        def bulk_classifications():
            for charity_id, fyend, classification in tqdm(
                self.classifications, desc="Processing classifications"
            ):
                yield CharityFinancialYear.classifications.through(
                    charityfinancialyear_id=lookup[(charity_id, fyend)],
                    charityclassificationcategory=classification,
                )

        CharityFinancialYear.classifications.through.objects.bulk_create(
            bulk_classifications(),
            batch_size=self.page_size,
            ignore_conflicts=True,
        )
        self.logger("Finished importing annual report classifications.")

    def extract_website(self, text: str) -> str | None:
        import re

        # Simple regex to find URLs in the text
        url_pattern = re.compile(r"(https?://[^\s]+)")
        match = url_pattern.search(text)
        if match:
            return match.group(0)
        return None

    def update_latest_financial_years(self) -> None:
        # Update the latest financial years for all charities
        latest_fy = (
            CharityFinancialYear.objects.values("charity_id")
            .annotate(max_fyend=models.Max("period_end_date"))
            .values("charity_id", "max_fyend")
        )
        classifications_to_update = CharityClassificationCategory.objects.filter(
            classification_type__in=[
                ClassificationTypes.REPORT_ACTIVITY,
                ClassificationTypes.BENEFICIARIES,
            ]
        ).values_list("id", flat=True)
        for latest_fy in tqdm(latest_fy, desc="Updating latest financial years"):
            latest = CharityFinancialYear.objects.get(
                charity_id=latest_fy["charity_id"],
                period_end_date=latest_fy["max_fyend"],
            )

            updates = dict(
                latest_financial_year_end=latest_fy["max_fyend"],
                latest_income=latest.gross_income,
                latest_expenditure=latest.gross_expenditure,
                latest_activity_description=latest.activity_description,
                latest_activity_description_ga=latest.activity_description_ga,
            )

            # look for a website in the latest_activity_description field
            website = self.extract_website(latest.activity_description)
            if website:
                updates["website"] = website

            Charity.objects.filter(
                registered_charity_number=latest_fy["charity_id"]
            ).update(**updates)
            Charity.classifications.through.objects.filter(
                charityclassificationcategory_id__in=classifications_to_update
            ).delete()
            Charity.classifications.through.objects.bulk_create(
                [
                    Charity.classifications.through(
                        charityclassificationcategory_id=id,
                        charity_id=latest.charity_id,
                    )
                    for id in latest.classifications.values_list("id", flat=True)
                ]
            )
